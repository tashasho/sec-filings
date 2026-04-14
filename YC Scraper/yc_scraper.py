#!/usr/bin/env python3
"""
YC Scraper: Scrapes Y Combinator companies & founders from W23–W25 batches.
Identifies Indian-origin founders (bachelor's in India) and highlights IIT alums.
Outputs a multi-tab Excel file (.xlsx) for Google Sheets import.
"""

import json
import time
import re
import html
import requests
import os
import sys
from pathlib import Path
from datetime import datetime

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuration ───────────────────────────────────────────────────────────

TARGET_BATCHES = ["W23", "S23", "W24", "S24", "W25", "Winter 2025", "Spring 2025", "Summer 2025", "Fall 2025", "Winter 2026", "Spring 2026", "Summer 2026"]
YC_OSS_API_BASE = "https://yc-oss.github.io/api/batches"

OUTPUT_DIR = Path(__file__).parent
OUTPUT_FILE = OUTPUT_DIR / "yc_indian_founders_iit_alums.xlsx"
DATA_DIR = OUTPUT_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

REQUEST_DELAY = 0.1  # seconds between requests to YC pages
MAX_RETRIES = 3

# ─── IIT Campuses ────────────────────────────────────────────────────────────

IIT_CAMPUSES = [
    "IIT Bombay", "IIT Delhi", "IIT Madras", "IIT Kanpur", "IIT Kharagpur",
    "IIT Roorkee", "IIT Guwahati", "IIT Hyderabad", "IIT Indore",
    "IIT BHU", "IIT (BHU)", "IIT Varanasi",
    "IIT Ropar", "IIT Patna", "IIT Bhubaneswar", "IIT Gandhinagar",
    "IIT Jodhpur", "IIT Mandi", "IIT Tirupati", "IIT Palakkad",
    "IIT Dharwad", "IIT Bhilai", "IIT Goa", "IIT Jammu",
    "IIT (ISM) Dhanbad", "IIT ISM", "IIT Dhanbad",
    "Indian Institute of Technology",
]

# Regex-friendly campus name list for extraction
IIT_CAMPUS_NAMES = [
    "Bombay", "Delhi", "Madras", "Kanpur", "Kharagpur",
    "Roorkee", "Guwahati", "Hyderabad", "Indore",
    "BHU", "Varanasi", "Ropar", "Patna", "Bhubaneswar",
    "Gandhinagar", "Jodhpur", "Mandi", "Tirupati", "Palakkad",
    "Dharwad", "Bhilai", "Goa", "Jammu", "Dhanbad", "ISM",
]

# ─── Indian Institution Keywords ────────────────────────────────────────────

INDIAN_INSTITUTIONS = [
    # IITs
    "IIT", "Indian Institute of Technology",
    # NITs
    "NIT ", "National Institute of Technology", "NIT-", "NIT,",
    # IIITs
    "IIIT", "International Institute of Information Technology",
    "Indian Institute of Information Technology",
    # IISc, IISERs
    "IISc", "Indian Institute of Science",
    "IISER",
    # Top Universities
    "BITS Pilani", "BITS ", "Birla Institute of Technology",
    "Delhi University", "University of Delhi",
    "Jawaharlal Nehru University", "JNU",
    "Anna University",
    "VIT Vellore", "Vellore Institute of Technology",
    "SRM University", "SRM Institute",
    "Manipal Institute", "Manipal University", "MIT Manipal",
    "NSIT", "Netaji Subhas Institute",
    "DTU", "Delhi Technological University", "Delhi College of Engineering",
    "Jadavpur University",
    "PES University", "PESIT",
    "Thapar University", "Thapar Institute",
    "COEP", "College of Engineering, Pune", "College of Engineering Pune",
    "Mumbai University", "University of Mumbai",
    "Pune University", "Savitribai Phule",
    "Bangalore University",
    "Osmania University",
    "DAIICT", "DA-IICT",
    "Shiv Nadar University",
    "Ashoka University",
    "ISI Kolkata", "Indian Statistical Institute",
    "VJTI",
    "ICT Mumbai",
    "College of Engineering Guindy",
    "Motilal Nehru National Institute", "MNNIT",
    "BIT Mesra",
    "ISM Dhanbad",
    "Harcourt Butler",
    "IIITM Gwalior",
    "LNM Institute", "LNMIIT",
    "PSG College", "PSG Tech",
    "RV College", "RVCE",
    "BMS College", "BMSCE",
    "Punjab Engineering College", "PEC Chandigarh",
    "St. Stephen", "St Stephen",
    "SRCC", "Shri Ram College",
    "Hindu College",
    "Hansraj College",
    "Loyola College",
    "Presidency University", "Presidency College",
    "Calcutta University", "University of Calcutta",
    "Madras University", "University of Madras",
    "Amity University",
    "Christ University",
    "Symbiosis",
    "NMIMS",
    "XLRI",
    "IIM ", "Indian Institute of Management",
    "ISB ", "Indian School of Business",
    "SP Jain",
    "MDI Gurgaon",
    "FMS Delhi",
    "JNTU", "Jawaharlal Nehru Technological University",
    "Visvesvaraya",
    "Dhirubhai Ambani", "DAIICT",
    "KIIT", "Kalinga Institute",
    "Manipal Academy",
    "SRM", 
    "VIT ",
    "SRMIST",
    "Birla",
    "IIST",  # Indian Institute of Space Science
]

# Indian degree patterns (for bachelor's detection)
INDIAN_DEGREE_PATTERNS = [
    r"B\.?\s*Tech",
    r"B\.?\s*E\.?(?:\s|,|\()",
    r"Bachelor.*(?:Engineering|Technology|Science|Computer|Electrical|Mechanical|Chemical|Civil|Electronics)",
    r"BTech",
    r"BE\b",
    r"B\.Sc",
    r"B\.A\b",
    r"BCA\b",
    r"BBA\b",
    r"undergraduate",
]


# ─── Step 1: Fetch Companies from yc-oss API ────────────────────────────────

def fetch_companies_from_ycoss(batch: str) -> list:
    """Fetch all companies for a batch from yc-oss community API."""
    batch_lower = batch.lower()
    url = f"{YC_OSS_API_BASE}/{batch_lower}.json"
    
    try:
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            print(f"  ✅ {batch}: {len(data)} companies")
            return data
        elif resp.status_code == 404:
            print(f"  ⚠️  {batch}: Not available yet (404)")
            return []
        else:
            print(f"  ⚠️  {batch}: HTTP {resp.status_code}")
            return []
    except Exception as e:
        print(f"  ❌ {batch}: Error — {e}")
        return []


def run_step1_fetch_companies() -> dict:
    """Step 1: Fetch all companies across all target batches."""
    print("=" * 70)
    print("STEP 1: Fetching company lists from yc-oss API")
    print("=" * 70)
    
    all_companies = {}
    
    for batch in TARGET_BATCHES:
        print(f"\n📦 Batch {batch}:")
        companies = fetch_companies_from_ycoss(batch)
        
        for c in companies:
            slug = c.get("slug", "")
            if not slug:
                continue
            all_companies[slug] = {
                "name": c.get("name", ""),
                "slug": slug,
                "batch": c.get("batch", batch),
                "one_liner": c.get("one_liner", ""),
                "long_description": c.get("long_description", ""),
                "website": c.get("website", ""),
                "all_locations": c.get("all_locations", ""),
                "team_size": c.get("team_size", ""),
                "industry": c.get("industry", ""),
                "subindustry": c.get("subindustry", ""),
                "status": c.get("status", ""),
                "stage": c.get("stage", ""),
                "tags": ", ".join(c.get("tags", [])) if isinstance(c.get("tags"), list) else str(c.get("tags", "")),
                "top_company": c.get("top_company", False),
                "isHiring": c.get("isHiring", False),
                "url": c.get("url", f"https://www.ycombinator.com/companies/{slug}"),
                "logo": c.get("small_logo_thumb_url", ""),
                "founders": [],
            }
    
    # Load externally scraped new batches
    playwright_file = DATA_DIR / "playwright_companies.json"
    if playwright_file.exists():
        with open(playwright_file) as f:
            playwright_companies = json.load(f)
            for slug, comp in playwright_companies.items():
                all_companies[slug] = comp
        print(f"\n  ✅ Loaded {len(playwright_companies)} companies from playwright scrape.")

    print(f"\n{'─' * 50}")
    print(f"✅ Total companies across all batches: {len(all_companies)}")
    
    # Save
    with open(DATA_DIR / "step1_companies.json", "w") as f:
        json.dump(all_companies, f, indent=2)
    
    return all_companies


# ─── Step 2: Fetch Founder Details from YC Pages ────────────────────────────

def fetch_yc_page(slug: str, attempt: int = 1) -> str | None:
    """Fetch the HTML of a YC company page with retry logic."""
    url = f"https://www.ycombinator.com/companies/{slug}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        if resp.status_code == 200:
            return resp.text
        elif resp.status_code == 429 and attempt <= MAX_RETRIES:
            wait = 5 * attempt
            print(f"    ⏳ Rate limited on {slug}, waiting {wait}s...")
            time.sleep(wait)
            return fetch_yc_page(slug, attempt + 1)
        else:
            return None
    except Exception:
        if attempt <= MAX_RETRIES:
            time.sleep(2)
            return fetch_yc_page(slug, attempt + 1)
        return None


def extract_founders_from_page(page_html: str) -> list:
    """Extract founder data from YC company page using Inertia.js data-page JSON."""
    soup = BeautifulSoup(page_html, "html.parser")
    founders = []
    
    # Method 1: Look for data-page attribute (Inertia.js)
    page_div = soup.find(attrs={"data-page": True})
    if page_div:
        try:
            raw = page_div["data-page"]
            # Decode HTML entities
            decoded = html.unescape(raw)
            page_data = json.loads(decoded)
            
            props = page_data.get("props", {})
            company = props.get("company", {})
            
            for f in company.get("founders", []):
                founders.append({
                    "name": (f.get("full_name") or "").strip(),
                    "bio": (f.get("founder_bio") or "").strip(),
                    "title": (f.get("title") or "Founder").strip(),
                    "linkedin": (f.get("linkedin_url") or "").strip(),
                    "twitter": (f.get("twitter_url") or "").strip(),
                    "avatar": f.get("avatar_thumb_url") or "",
                })
        except (json.JSONDecodeError, KeyError, TypeError):
            pass
    
    # Method 2: Fallback — look for JSON in script tags
    if not founders:
        for script in soup.find_all("script"):
            text = script.string or ""
            if "founders" in text and "full_name" in text:
                try:
                    # Try to find JSON object
                    match = re.search(r'\{[^{}]*"founders"\s*:\s*\[.*?\]\s*[^{}]*\}', text, re.DOTALL)
                    if match:
                        data = json.loads(match.group())
                        for f in data.get("founders", []):
                            founders.append({
                                "name": (f.get("full_name") or "").strip(),
                                "bio": (f.get("founder_bio") or "").strip(),
                                "title": (f.get("title") or "Founder").strip(),
                                "linkedin": (f.get("linkedin_url") or "").strip(),
                                "twitter": (f.get("twitter_url") or "").strip(),
                                "avatar": f.get("avatar_thumb_url") or "",
                            })
                except:
                    pass
    
    return founders


def run_step2_fetch_founders(all_companies: dict) -> dict:
    """Step 2: Fetch founder details from each company's YC page."""
    print("\n" + "=" * 70)
    print("STEP 2: Scraping founder details from YC company pages")
    print("=" * 70)
    
    total = len(all_companies)
    found_count = 0
    error_count = 0
    total_founders = 0
    
    # Check for checkpoint
    checkpoint_file = DATA_DIR / "step2_checkpoint.json"
    if checkpoint_file.exists():
        print("  📂 Found checkpoint, resuming...")
        with open(checkpoint_file) as f:
            checkpoint = json.load(f)
        for slug, founders in checkpoint.items():
            if slug in all_companies:
                all_companies[slug]["founders"] = founders
                if founders:
                    found_count += 1
                    total_founders += len(founders)
        print(f"  ✅ Loaded {found_count} companies from checkpoint")
    
    slugs = list(all_companies.keys())
    checkpoint_data = {}
    
    for i, slug in enumerate(slugs):
        company = all_companies[slug]
        
        # Skip if already have founders from checkpoint
        if company.get("founders"):
            checkpoint_data[slug] = company["founders"]
            continue
        
        # Progress indicator
        if (i + 1) % 25 == 0 or i == 0:
            print(f"\n  📊 Progress: {i + 1}/{total} companies "
                  f"({found_count} with founders, {total_founders} total founders)")
        
        page_html = fetch_yc_page(slug)
        if page_html:
            founders = extract_founders_from_page(page_html)
            if founders:
                company["founders"] = founders
                found_count += 1
                total_founders += len(founders)
                checkpoint_data[slug] = founders
            else:
                error_count += 1
                checkpoint_data[slug] = []
        else:
            error_count += 1
            checkpoint_data[slug] = []
        
        # Save checkpoint every 50 companies
        if (i + 1) % 50 == 0:
            with open(checkpoint_file, "w") as f:
                json.dump(checkpoint_data, f)
            print(f"  💾 Checkpoint saved ({i + 1}/{total})")
        
        time.sleep(REQUEST_DELAY)
    
    # Final checkpoint
    with open(checkpoint_file, "w") as f:
        json.dump(checkpoint_data, f)
    
    print(f"\n{'─' * 50}")
    print(f"✅ Founder scraping complete:")
    print(f"   Companies with founders: {found_count}/{total}")
    print(f"   Total founders found: {total_founders}")
    print(f"   Errors/missing: {error_count}")
    
    # Save full data
    with open(DATA_DIR / "step2_companies_with_founders.json", "w") as f:
        json.dump(all_companies, f, indent=2)
    
    return all_companies


# ─── Step 3: Detect Indian-Origin & IIT Alumni ──────────────────────────────

def detect_iit(text: str) -> tuple:
    """
    Detect if text mentions any IIT.
    Returns (is_iit: bool, campus: str).
    """
    if not text:
        return False, ""
    
    # Check for specific IIT campus mentions
    for campus in IIT_CAMPUSES:
        if campus.lower() in text.lower():
            return True, campus
    
    # Check for "IIT" with campus name nearby
    iit_match = re.search(
        r'\bIIT[\s\-,]*(' + '|'.join(IIT_CAMPUS_NAMES) + r')\b',
        text, re.IGNORECASE
    )
    if iit_match:
        return True, f"IIT {iit_match.group(1).strip()}"
    
    # Generic IIT mention
    if re.search(r'\bIIT\b', text) and not re.search(r'\bIITM\b|\bIITR\b', text):
        # Try to extract campus from surrounding context
        campus_match = re.search(
            r'(?:IIT|Indian Institute of Technology)\s*[,\-\(]?\s*(' + 
            '|'.join(IIT_CAMPUS_NAMES) + r')',
            text, re.IGNORECASE
        )
        campus = f"IIT {campus_match.group(1)}" if campus_match else "IIT (campus not specified)"
        return True, campus
    
    # "Indian Institute of Technology" spelled out
    if "indian institute of technology" in text.lower():
        campus_match = re.search(
            r'Indian Institute of Technology\s*[,\-\(]?\s*(' + 
            '|'.join(IIT_CAMPUS_NAMES) + r')',
            text, re.IGNORECASE
        )
        campus = f"IIT {campus_match.group(1)}" if campus_match else "IIT (campus not specified)"
        return True, campus
    
    return False, ""


def detect_indian_bachelor(text: str) -> bool:
    """Check if text indicates a bachelor's degree from an Indian institution."""
    if not text:
        return False
    
    text_lower = text.lower()
    
    # Check for Indian institution mentions
    for inst in INDIAN_INSTITUTIONS:
        if inst.lower() in text_lower:
            return True
    
    # Check for IIT
    is_iit_flag, _ = detect_iit(text)
    if is_iit_flag:
        return True
    
    return False


def extract_education_details(bio: str) -> dict:
    """
    Extract education details from a founder bio.
    Returns dict with keys: university, degree, grad_year, is_iit, iit_campus.
    """
    result = {
        "university": "",
        "degree": "",
        "grad_year": "",
        "is_iit": False,
        "iit_campus": "",
    }
    
    if not bio:
        return result
    
    # Detect IIT
    is_iit_flag, campus = detect_iit(bio)
    result["is_iit"] = is_iit_flag
    result["iit_campus"] = campus
    if is_iit_flag:
        result["university"] = campus
    
    # If no IIT, try to find other Indian institutions
    if not result["university"]:
        for inst in INDIAN_INSTITUTIONS:
            if inst.lower() in bio.lower():
                # Try to get full university name from context
                pattern = re.escape(inst) + r'[\w\s,\-]*'
                match = re.search(pattern, bio, re.IGNORECASE)
                if match:
                    result["university"] = match.group().strip().rstrip(",.- ")
                else:
                    result["university"] = inst
                break
    
    # Extract degree/course
    degree_patterns = [
        # B.Tech / BTech
        r"(B\.?\s*Tech\.?(?:\s+(?:in\s+)?[\w\s&]+)?)",
        # B.E.
        r"(B\.?\s*E\.?\s+(?:in\s+)?[\w\s&]+)",
        # Bachelor of/in
        r"(Bachelor(?:'s)?\s+(?:of|in)\s+[\w\s&]+)",
        # Specific degree names
        r"((?:Computer Science|Electrical Engineering|Mechanical Engineering|"
        r"Chemical Engineering|Civil Engineering|Electronics|"
        r"Computer Engineering|Information Technology|"
        r"Aerospace Engineering|Metallurgical|Mining|Biotechnology|"
        r"Physics|Chemistry|Mathematics|Economics)\s*(?:and\s+[\w\s]+)?)",
        # B.Sc/BSc
        r"(B\.?\s*Sc\.?\s+(?:in\s+)?[\w\s&]+)",
        # MS/M.Tech (for completeness)
        r"(M\.?\s*Tech\.?(?:\s+(?:in\s+)?[\w\s&]+)?)",
        r"(M\.?\s*S\.?\s+(?:in\s+)?[\w\s&]+)",
        r"(Master(?:'s)?\s+(?:of|in)\s+[\w\s&]+)",
    ]
    
    for pattern in degree_patterns:
        match = re.search(pattern, bio, re.IGNORECASE)
        if match:
            degree = match.group(1).strip()
            # Clean up
            degree = re.sub(r'\s+', ' ', degree)
            degree = degree.rstrip(",.- ")
            if len(degree) > 5:  # Avoid tiny matches
                result["degree"] = degree
                break
    
    # Extract graduation year
    year_patterns = [
        r"(?:graduated?|class of|batch of|\')\s*(\d{4})",
        r"(\d{4})\s*(?:graduate|batch|class|alumnus|alumna|alumni)",
        r"(?:in|from)\s+(\d{4})",
        r"(?:19|20)(\d{2})\s*[-–]\s*(?:19|20)(\d{2})",  # Range like 2012-2016
    ]
    
    for pattern in year_patterns:
        match = re.search(pattern, bio, re.IGNORECASE)
        if match:
            groups = match.groups()
            if len(groups) == 2:
                # Year range — take the graduation (end) year
                result["grad_year"] = f"20{groups[1]}" if len(groups[1]) == 2 else groups[1]
            else:
                year = int(groups[0])
                if 1990 <= year <= 2030:
                    result["grad_year"] = str(year)
            break
    
    # Second pass for standalone years near education context
    if not result["grad_year"] and result["university"]:
        years = re.findall(r'\b((?:19|20)\d{2})\b', bio)
        for y in years:
            y_int = int(y)
            if 1990 <= y_int <= 2026:
                result["grad_year"] = y
                break
    
    return result


def run_step3_screen_founders(all_companies: dict) -> tuple:
    """Step 3: Screen founders for Indian origin and IIT background."""
    print("\n" + "=" * 70)
    print("STEP 3: Screening for Indian-origin founders & IIT alumni")
    print("=" * 70)
    
    indian_founders = []
    iit_founders = []
    
    for slug, company in all_companies.items():
        for founder in company.get("founders", []):
            bio = founder.get("bio", "")
            long_desc = company.get("long_description", "")
            
            # Combine bio and company description for searching
            search_text = f"{bio} {long_desc}"
            
            # Check for Indian bachelor's
            if not detect_indian_bachelor(search_text):
                continue
            
            # Extract education details
            edu = extract_education_details(search_text)
            
            record = {
                "name": founder.get("name", ""),
                "bio": bio,
                "company": company.get("name", ""),
                "batch": company.get("batch", ""),
                "title": founder.get("title", ""),
                "linkedin": founder.get("linkedin", ""),
                "twitter": founder.get("twitter", ""),
                "university": edu["university"],
                "degree": edu["degree"],
                "grad_year": edu["grad_year"],
                "is_iit": edu["is_iit"],
                "iit_campus": edu["iit_campus"],
                "company_website": company.get("website", ""),
                "company_industry": company.get("industry", ""),
                "company_subindustry": company.get("subindustry", ""),
                "company_one_liner": company.get("one_liner", ""),
                "company_status": company.get("status", ""),
                "company_stage": company.get("stage", ""),
                "team_size": company.get("team_size", ""),
                "location": company.get("all_locations", ""),
                "tags": company.get("tags", ""),
                "yc_url": company.get("url", ""),
            }
            
            indian_founders.append(record)
            
            if edu["is_iit"]:
                iit_founders.append(record)
    
    # Sort by batch
    batch_order = {b: i for i, b in enumerate(TARGET_BATCHES)}
    indian_founders.sort(key=lambda x: batch_order.get(x["batch"], 99))
    iit_founders.sort(key=lambda x: batch_order.get(x["batch"], 99))
    
    print(f"\n{'─' * 50}")
    print(f"✅ Screening results:")
    print(f"   Indian-origin founders (bachelor's in India): {len(indian_founders)}")
    print(f"   IIT alumni: {len(iit_founders)}")
    
    # Batch breakdown
    for batch in TARGET_BATCHES:
        indian_in_batch = sum(1 for f in indian_founders if f["batch"] == batch)
        iit_in_batch = sum(1 for f in iit_founders if f["batch"] == batch)
        if indian_in_batch > 0:
            print(f"   {batch}: {indian_in_batch} Indian-origin, {iit_in_batch} IIT")
    
    # Save
    with open(DATA_DIR / "step3_indian_founders.json", "w") as f:
        json.dump(indian_founders, f, indent=2)
    with open(DATA_DIR / "step3_iit_founders.json", "w") as f:
        json.dump(iit_founders, f, indent=2)
    
    return indian_founders, iit_founders


# ─── Step 4: Generate Multi-Tab Excel ───────────────────────────────────────

def generate_excel(all_companies: dict, indian_founders: list, iit_founders: list):
    """Generate a professional multi-tab Excel workbook."""
    print("\n" + "=" * 70)
    print("STEP 4: Generating Excel workbook")
    print("=" * 70)
    
    wb = Workbook()
    
    # ─── Styles ──────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Calibri", size=10)
    
    fill_orange = PatternFill(start_color="FF6633", end_color="FF6633", fill_type="solid")
    fill_blue = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    fill_green = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
    fill_gold = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    fill_purple = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    
    iit_row_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(vertical="top", wrap_text=True)
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    
    def style_header(ws, headers, fill):
        ws.freeze_panes = "A2"  # Freeze header row
        for col_num, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_text)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = header_align
            cell.border = thin_border
    
    def auto_width(ws, max_width=45):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_len = len(str(cell.value or ""))
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 3, max_width)
    
    def add_data_row(ws, row_data, row_num, highlight=False):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if highlight:
                cell.fill = iit_row_fill
            elif row_num % 2 == 0:
                cell.fill = alt_row_fill
    
    batch_order = {b: i for i, b in enumerate(TARGET_BATCHES)}
    
    # ─── Tab 1: All Companies ────────────────────────────────────
    print("  📋 Creating 'All Companies' tab...")
    ws1 = wb.active
    ws1.title = "All Companies"
    headers1 = [
        "Company Name", "Batch", "One-Liner", "Industry", "Sub-Industry",
        "Website", "Location", "Team Size", "Status", "Stage",
        "Tags", "Top Company", "Hiring?", "# Founders", "Founder Names", "YC URL"
    ]
    style_header(ws1, headers1, fill_orange)
    
    sorted_companies = sorted(all_companies.values(), key=lambda x: batch_order.get(x["batch"], 99))
    for i, c in enumerate(sorted_companies):
        founder_names = ", ".join([f.get("name", "") for f in c.get("founders", []) if f.get("name")])
        row = [
            c["name"], c["batch"], c["one_liner"], c["industry"], c["subindustry"],
            c["website"], c["all_locations"], c["team_size"], c["status"], c["stage"],
            c["tags"], "Yes" if c["top_company"] else "No",
            "Yes" if c["isHiring"] else "No", len(c.get("founders", [])),
            founder_names, c["url"]
        ]
        add_data_row(ws1, row, i + 2)
    auto_width(ws1)
    
    # ─── Tab 2: All Founders ─────────────────────────────────────
    print("  👤 Creating 'All Founders' tab...")
    ws2 = wb.create_sheet("All Founders")
    headers2 = [
        "Founder Name", "Company", "Batch", "Title/Role",
        "Founder Bio", "LinkedIn", "Twitter",
        "Company Website", "Company Industry", "Company One-Liner"
    ]
    style_header(ws2, headers2, fill_blue)
    
    row_num = 2
    for c in sorted_companies:
        for f in c.get("founders", []):
            row = [
                f.get("name", ""), c["name"], c["batch"], f.get("title", ""),
                f.get("bio", ""), f.get("linkedin", ""), f.get("twitter", ""),
                c["website"], c["industry"], c["one_liner"]
            ]
            add_data_row(ws2, row, row_num)
            row_num += 1
    auto_width(ws2)
    
    # ─── Tab 3: Indian-Origin Founders ───────────────────────────
    print("  🇮🇳 Creating 'Indian-Origin Founders' tab...")
    ws3 = wb.create_sheet("Indian-Origin Founders")
    headers3 = [
        "Founder Name", "Company", "Batch", "University (Bachelor's)",
        "Degree/Course", "Graduation Year", "IIT Alum?", "IIT Campus",
        "LinkedIn", "Founder Bio", "Title/Role",
        "Company Website", "Company Industry", "Company One-Liner", "YC URL"
    ]
    style_header(ws3, headers3, fill_green)
    
    for i, f in enumerate(indian_founders):
        is_iit = f.get("is_iit", False)
        row = [
            f["name"], f["company"], f["batch"], f["university"],
            f["degree"], f["grad_year"],
            f"✅ Yes — {f['iit_campus']}" if is_iit else "No",
            f["iit_campus"] if is_iit else "",
            f["linkedin"], f["bio"], f["title"],
            f["company_website"], f["company_industry"],
            f["company_one_liner"], f["yc_url"]
        ]
        add_data_row(ws3, row, i + 2, highlight=is_iit)
    auto_width(ws3)
    
    # ─── Tab 4: IIT Alums ⭐ ─────────────────────────────────────
    print("  ⭐ Creating 'IIT Alums' tab...")
    ws4 = wb.create_sheet("IIT Alums ⭐")
    headers4 = [
        "Founder Name", "Company", "Batch/Cohort", "IIT Campus",
        "Degree/Course", "Graduation Year", "Founder Bio",
        "LinkedIn", "Title/Role",
        "Company One-Liner", "Company Industry", "Sub-Industry",
        "Company Website", "Company Status", "Stage",
        "Team Size", "Location", "Tags", "YC URL"
    ]
    style_header(ws4, headers4, fill_gold)
    
    for i, f in enumerate(iit_founders):
        row = [
            f["name"], f["company"], f["batch"], f["iit_campus"],
            f["degree"], f["grad_year"], f["bio"],
            f["linkedin"], f["title"],
            f["company_one_liner"], f["company_industry"], f["company_subindustry"],
            f["company_website"], f["company_status"], f["company_stage"],
            f["team_size"], f["location"], f["tags"], f["yc_url"]
        ]
        add_data_row(ws4, row, i + 2, highlight=True)
    auto_width(ws4)
    
    # ─── Tab 5: Summary & Stats ──────────────────────────────────
    print("  📊 Creating 'Summary' tab...")
    ws5 = wb.create_sheet("Summary & Stats")
    headers5 = ["Metric", "Value"]
    style_header(ws5, headers5, fill_purple)
    
    total_founders = sum(len(c.get("founders", [])) for c in all_companies.values())
    
    stats = [
        ("", ""),
        ("📋 OVERALL", ""),
        ("Total Companies (W23–W25)", len(all_companies)),
        ("Total Founders Scraped", total_founders),
        ("Indian-Origin Founders (Bachelor's in India)", len(indian_founders)),
        ("IIT Alumni", len(iit_founders)),
        ("", ""),
        ("📦 BATCH-WISE BREAKDOWN", ""),
    ]
    
    for batch in TARGET_BATCHES:
        companies_in_batch = sum(1 for c in all_companies.values() if c["batch"] == batch)
        if companies_in_batch > 0:
            founders_in_batch = sum(
                len(c.get("founders", [])) for c in all_companies.values() if c["batch"] == batch
            )
            indian_in_batch = sum(1 for f in indian_founders if f["batch"] == batch)
            iit_in_batch = sum(1 for f in iit_founders if f["batch"] == batch)
            stats.append((f"", ""))
            stats.append((f"--- {batch} ---", ""))
            stats.append((f"  Companies", companies_in_batch))
            stats.append((f"  Founders", founders_in_batch))
            stats.append((f"  Indian-Origin Founders", indian_in_batch))
            stats.append((f"  IIT Alumni", iit_in_batch))
    
    # IIT Campus breakdown
    if iit_founders:
        stats.append(("", ""))
        stats.append(("🏫 IIT CAMPUS BREAKDOWN", ""))
        campus_counts = {}
        for f in iit_founders:
            campus = f.get("iit_campus", "Unknown")
            campus_counts[campus] = campus_counts.get(campus, 0) + 1
        for campus, count in sorted(campus_counts.items(), key=lambda x: -x[1]):
            stats.append((f"  {campus}", count))
    
    for i, (label, value) in enumerate(stats):
        row_num = i + 2
        cell_label = ws5.cell(row=row_num, column=1, value=label)
        cell_value = ws5.cell(row=row_num, column=2, value=value)
        cell_label.font = Font(name="Calibri", size=11, bold=("─" in label or "📋" in label or "📦" in label or "🏫" in label))
        cell_value.font = Font(name="Calibri", size=11, bold=True)
        cell_label.border = thin_border
        cell_value.border = thin_border
    
    ws5.column_dimensions["A"].width = 45
    ws5.column_dimensions["B"].width = 15
    
    # Save
    wb.save(OUTPUT_FILE)
    print(f"\n{'─' * 50}")
    print(f"✅ Excel file saved: {OUTPUT_FILE}")
    print(f"   → Import into Google Sheets: File → Import → Upload")
    
    return str(OUTPUT_FILE)


# ─── Main Entry Point ───────────────────────────────────────────────────────

def main():
    start_time = datetime.now()
    
    print("🚀 YC Scraper: Indian-Origin Founders & IIT Alumni")
    print(f"   Batches: {', '.join(TARGET_BATCHES)}")
    print(f"   Started: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    # Step 1: Fetch company lists
    all_companies = run_step1_fetch_companies()
    
    # Step 2: Fetch founder details from YC pages
    all_companies = run_step2_fetch_founders(all_companies)
    
    # Step 3: Screen for Indian-origin and IIT alumni
    indian_founders, iit_founders = run_step3_screen_founders(all_companies)
    
    # Step 4: Generate Excel
    output_path = generate_excel(all_companies, indian_founders, iit_founders)
    
    elapsed = datetime.now() - start_time
    print(f"\n{'═' * 70}")
    print(f"🎉 DONE! Total time: {elapsed}")
    print(f"   Output: {output_path}")
    print(f"{'═' * 70}")


if __name__ == "__main__":
    main()
